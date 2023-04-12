import pickle

'''
 - Need to check whether time is present and run if it is'''
import time
import pandas as pd
import json
from openpyxl import load_workbook
from copy import copy
#Lines 4~197 for input

#Read information to calculate from this file:
template = pd.read_excel("read.xlsx")
template = pd.DataFrame(template)



date = str(template.iloc[8][1]).split(' ')[0]

#Get numeric characters
def integerExtractor(var):
    integers = []
    inttemp = []
    #only get integers
    for x in var:
        try:
            x = int(x)
        except:
            pass
        else:
            integers.append(x)
    for a in integers:
        inttemp.append(str(a))
    output = ''.join(inttemp)
    integer = int(output)
    return integer
    
#Add colon for xx:xx if it is needed
def colonorNot(data):
    temp=[]
    for x in data:
        temp.append(str(x))
    output = ''.join(temp)
    
    if len(output)==4:
        output = f"{output[:2]}:{output[2:]}"
    elif len(output)==3:
        output = f"{output[:1]}:{output[1:]}"
    else:
        return
    return output
    
#Detect whether the time is am or pm based off user input 
def appendampm(x):
    if x is None:
        return
    if 'am' in x.lower():
        y = 'am'
    elif 'pm' in x.lower():
        y = 'pm'
    else:
        return None, x
    return y, str(x)
        
#Add zeros for time computation
def zeros(data):
    print(f'{data}, {len(data)}, {type(data)}')
    if len(str(data))==1:
        slice = 1
        temp = list(str(data))
        temp.insert(slice, '00')
        data = int(''.join(temp))
        return data
    elif len(str(data))==2:
        slice = 2
        temp = list(str(data))
        temp.insert(slice, '00')
        data = int(''.join(temp))
        return data
    else:
        return(int(data))


#Calculate all values
def calculate(position):
    endposition = position
    c = 0
    while c<=5:
        if str(template.iloc[position+c][6])!='nan':
            print(template.iloc[position+c][6])
            endposition = position+c+1
        else:
            c=5
        c=c+1

    print(position)
    print(endposition)
    #detecting whether each time period is am or pm
    dashindex = str(template.iloc[position][4]).index("-")
    ampm = ['', '']
    ampm[0], Start = appendampm(str(template.iloc[position][4][:dashindex]))
    print(ampm[0])
    ampm[1], End = appendampm(str(template.iloc[position][4][dashindex:]))

    #idnetify vehicle type based on Excel Form
    vehicleTypes=[]
    for i in range(position, endposition):
        if template.iloc[i][6]:
            vehicleTypes.append(f'{template.iloc[i][6]}')
            print(template.iloc[i][6])

    #Identify number of vehicles
    numVehicles = []
    for i in range(position,endposition):
        if template.iloc[i][5]:
            numVehicles.append(integerExtractor(str(template.iloc[i][5])))

    numDrivers = template.iloc[i][1]
    numSupervisors = template.iloc[i][3]


    #Identify rate based on Excel Form
    RateDict = {'Towing Package(s)': '561.84', 
    'Impact Suppression Vehicle(s)': '40', 
    'Motor Vehicle Operator(s)': '32',
    'Supervisor(s)': '38.75',
    'Tow Truck(s)': '83.22',
    'Tow Truck (PEMA)': '83.22',
    'Tow Truck Driver(s)': '37.59',
    'Parking Enforcement Vehicle(s)': '40',
    'Parking Enforcement Officer(s)': '35.09',
    'Impact Suppression Vehicle(s)': '40',
    'Dump Truck (SWMA-ISV)': '40',  
    'Flusher (SWMA)': '40', 
    'Trash Truck (SWMA)': '40',
    'Recycling Truck (SWMA)': '40', 
    'Dump Truck (SWMA)': '40', 
    'Sweeper(s)': '40',
    'Liftgate (SWMA)': '40', 
    'Dumpster (SWMA)': '40', 
    'Flatbed Truck (FMA)': '40', 
    'Flatbed Truck (PEMA)': '40', 
    'Large Tow (FMA)': '40', 
    'Large Tow (PEMA)': '40', 
    'Gator (FMA)': '40', 
    'Solar Light Tower (FMA)': '40'}
    RateArr = []
    for vehicleType in vehicleTypes:
        RateArr.append(RateDict[vehicleType])



    #Convert values:
    #xx00
    intStart = integerExtractor(Start)
    #xx00
    intEnd = integerExtractor(End)
    #xx00
    intRates = []
    for Rate in RateArr:
        intRates.append(float(Rate))
    #convert the start time into a computable number
    intStart = zeros(str(intStart))
    #convert the start time into xx:xx
    timeStart = colonorNot(str(intStart))


    #Calculate amt of time
    intEnd = zeros(str(intEnd))
    timeEnd = colonorNot(str(intEnd))


    if intEnd == 1200:
        intEnd = 0
    
    if intStart == 1200:
        Hours = -1
    else:
        Hours = 0
    if intEnd!=intStart:
        while intStart!=intEnd and Hours<20:
            Hours+=1
            intStart+=100
            if intStart>1100:
                intStart = 0
    else:
        Hours = 12

    #Dictionary template to fill 
    dictionaryArr = []
    TotalAmounts = []
    for i in range(len(intRates)): 
        totVehicles = round((numVehicles[i]*Hours)*intRates[i], 2)
        totSupervisors = round(float(RateDict['Supervisor(s)'])*Hours*numSupervisors, 2)
        totDrivers = round(numVehicles[i]*float(RateDict['Motor Vehicle Operator(s)'])*Hours, 2)
        VehicleDict = {
            "Description": f"{numVehicles[i]} {vehicleTypes[i]}, each for {Hours} hours",
            "Service Hours": f"{timeStart} {ampm[0]} - {timeEnd} {ampm[1]}",
            "Total Hours": f"{Hours*numVehicles[i]} hours",
            "Hourly Rate": f"{intRates[i]} dollars an hour",
            "Total Amount": f"{totVehicles} dollars"
        }

        dictionaryArr.append(VehicleDict)
    DriverDict = {
        "Description": f"{numDrivers} drivers, each for {Hours} hours",
        "Service Hours": f"{timeStart} {ampm[0]} - {timeEnd} {ampm[1]}",
        "Total Hours": f"{Hours*numDrivers} hours",
        "Hourly Rate": f"{float(RateDict['Motor Vehicle Operator(s)'])} dollars an hour",
        "Total Amount": f"{totDrivers} dollars"
    }
    SupervisorDict = {
        "Description": f"{numSupervisors} supervisor for {Hours} hours",
        "Service Hours": f"{timeStart} {ampm[0]} - {timeEnd} {ampm[1]}",
        "Total Hours": f"{Hours} hours",
        "Hourly Rate": f"{float(RateDict['Supervisor(s)'])*numSupervisors} dollars an hour",
        "Total Amount": f"{totSupervisors} dollars"
    }
    dictionaryArr.append(DriverDict)
    dictionaryArr.append(SupervisorDict)



    #df = pd.DataFrame(data = dict, index=[0])

    return dictionaryArr, round(totVehicles+totDrivers+totSupervisors, 2)


dfarr = []
#Sum all costs and get a sum for the receipt
totalvalues = []
#Set up a final dataframe to append all dataframes to
finaldf = pd.DataFrame(data=None)

#n is used for figuring out what rows to delete
n=0
if type(template.iloc[13][4]) is str:
    dict0, value = calculate(13)
    for i in dict0:
        dfarr.append(i)
        n=n+1
    totalvalues.append(value)


if type(template.iloc[17][4]) is str:
    dict0, value = calculate(17)
    for i in dict0:
        dfarr.append(i)
        n=n+1
    totalvalues.append(value)

if type(template.iloc[22][4]) is str:
    dict0, value = calculate(22)
    for i in dict0:
        dfarr.append(i)
        n=n+1
    totalvalues.append(value)



#Used for multiple calculations
#main()

#Summation of cost
values = sum(totalvalues)
print(values)

dict1 = {
    "Description": '',
    "Service Hours": '',
    "Total Hours": '',
    "Hourly Rate": '',
    "Total Amount": f'{values} dollars'
}
#Fills the rest of the useless entries with dataframes of same size but no content
dictfill = {
    "Description": '',
    "Service Hours": '',
    "Total Hours": '',
    "Hourly Rate": '',
    "Total Amount": ''
}

summation = pd.DataFrame(data = dict1, index=[0])
dictfill = pd.DataFrame(data = dictfill, index=[0])

while len(dfarr)<20:
    dfarr.append(dictfill)
dfarr.append(summation)

print(len(dfarr))

#Copy is the template of the receipt that is used
path = pd.read_excel('writeto.xlsx')   
path = pd.DataFrame(path)

for i in range(19,40):
    if str(dfarr[i-19]["Total Amount"])[0]!='0':
        path.iloc[i][1] = str(dfarr[i-19]["Description"])
        path.iloc[i][2] = str(dfarr[i-19]["Service Hours"])
        path.iloc[i][3] = str(dfarr[i-19]["Total Hours"])
        path.iloc[i][4] = str(dfarr[i-19]["Hourly Rate"])
        path.iloc[i][5] = str(dfarr[i-19]["Total Amount"])
    elif i==39:
        path.iloc[i][5] = str(dfarr[i-19]["Total Amount"]).strip('0').strip('\nName: Total Amount, dtype: object')
    else:
        path.iloc[i][1] = ''
        path.iloc[i][2] = ''
        path.iloc[i][3] = ''
        path.iloc[i][4] = ''
        path.iloc[i][5] = ''
path.fillna(' ', inplace=True)
path.replace('\n', '', regex=True)
print(path)


style_attrs = ["alignment", "border", "fill", "font", "number_format", "protection"]

#Copy cell styling
def cells(worksheet):
    """Return a generator for the sequence of cells in the worksheet"""
    for row in worksheet:
        for cell in row:
            yield cell


def copy_attrs(src, dst, attrs=style_attrs):
    """Copy attributes from src to dst. Attributes are shallow-copied to avoid
    TypeError: unhashable type: 'StyleProxy'"""
    for name in attrs:
        setattr(dst, name, copy(getattr(src, name)))


def copy_column_attrs(worksheet_src, worksheet_dst, attrs=style_attrs + ["width"]):
    """Copy ColumnDimension properties from worksheet_src to worksheet_dst.
    Only properties listed in attrs will be copied."""
    for column, dimensions in worksheet_src.column_dimensions.items():
        copy_attrs(
            src=dimensions,
            dst=worksheet_dst.column_dimensions[column],
            attrs=style_attrs + ["width"],
        )



def copy_row_attrs(worksheet_src, worksheet_dst, attrs=style_attrs + ["height"]):
    """Copy RowDimension properties from worksheet_src to worksheet_dst.
    Only properties listed in attrs will be copied."""
    for row, dimensions in worksheet_src.row_dimensions.items():
        copy_attrs(
            src=dimensions,
            dst=worksheet_dst.row_dimensions[row],
            attrs=style_attrs + ["height"],
        )
num=0

#write new data from df information
def write_new_data(worksheet_dst, data):
    for i in range(19, 42):
        worksheet_dst.cell(row=i, column=1, value=data.iloc[i-2][0])
        worksheet_dst.cell(row=i, column=2, value=data.iloc[i-2][1])
        worksheet_dst.cell(row=i, column=3, value=data.iloc[i-2][2])
        worksheet_dst.cell(row=i, column=4, value=data.iloc[i-2][3])
        worksheet_dst.cell(row=i, column=5, value=data.iloc[i-2][4])
        worksheet_dst.cell(row=i, column=6, value=data.iloc[i-2][5])

#we only copy cell styles here:
def copy_cells(worksheet_src, worksheet_dst, data, attrs=style_attrs):
    global num
    arr=[]
    """Copy cells from worksheet_src to worksheet_dst. If cells are styled
    then also copy the attributes listed in attrs."""
    cellvalues = cells(worksheet_src)
    for cell in cells(worksheet_src):
        arr.append(cell)
        cell_dst = worksheet_dst.cell(row=cell.row, column=cell.column)
        if cell.has_style:
            copy_attrs(cell, cell_dst, attrs=attrs)
        cell_dst.value = cell.value
        num=num+1
    write_new_data(worksheet_dst, data)

#Clean end product before populating with data
def delete_worksheet_cells(worksheet):
    worksheet.delete_cols(1, worksheet.max_column + 1)
    worksheet.delete_rows(1, worksheet.max_row + 1)

#Template to copy style from
wb_src = load_workbook("source.xlsx")
ws_src = wb_src.active

#Document to write to with new data
wb_dst = load_workbook("writeto.xlsx")
ws_dst = wb_dst.active

delete_worksheet_cells(ws_dst)
copy_column_attrs(ws_src, ws_dst)
copy_row_attrs(ws_src, ws_dst)
#copy_cells and the functions used inside of it are the main problems
copy_cells(ws_src, ws_dst, path)


#Auto fit columns
dims = {}
for row in ws_dst.rows:
    for cell in row:
        if cell.value:
            dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
for col, value in dims.items():
    ws_dst.column_dimensions[col].width = 30

lis = []
f=[]
for col in range(48):
    lis.append(ws_dst.row_dimensions[col].height)
for i in lis:
    if i is not None:
        f.append(i)
ws_dst.row_dimensions[2].height=max(f)

#n was referenced when adding dataframes
ws_dst.delete_rows(21+n,41-(21+n))

#Save as excel file

class MyClass():
    def __init__(self, param):
        self.param = param
 
def save_object(obj):
    try:
        with open("meta/data.pickle", "wb") as f:
            pickle.dump(obj, f, protocol=pickle.HIGHEST_PROTOCOL)
    except Exception as ex:
        print("Error during pickling object (Possibly unsupported):", ex)
        return None
 
def load_object(filename):
    try:
        with open(filename, "rb") as f:
            return pickle.load(f)
    except Exception as ex:
        print("Error during unpickling object (Possibly unsupported):", ex)
        return None


#Check whether .pickle contains data
obj = int(load_object("meta/data.pickle"))
if obj == None:
    save_object(1)
    wb_dst.save(f'Outputs/output{str(obj).zfill(4)}.xlsx')
    obj=obj+1
    save_object(obj)
else:  
    wb_dst.save(f'Outputs/output{str(obj).zfill(4)}.xlsx')
    obj=obj+1
    save_object(obj)