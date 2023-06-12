import pickle
import time
import pandas as pd
import json
from openpyxl import load_workbook
from copy import copy
from timeconversion import convert

def main():


    #Lines 4~197 for input

    #Read information to calculate from this file:
    swma = pd.read_excel("swma.xlsx")
    swma = pd.DataFrame(swma)

    pema = pd.read_excel("pema.xlsx")
    pema = pd.DataFrame(pema)

    def LocateTimes(dat):
        TimeArrays = []
        for i in range(len(dat)):
            if '-' in str(dat.iloc[i][4]):
                print(i)
                print(dat.iloc[i][4])
                TimeArrays.append(i)
        return TimeArrays

    swmaTimes = LocateTimes(swma)
    pemaTimes = LocateTimes(pema)

    #Get numeric characters
    def integerExtractor(var):
        integers = []
        inttemp = []
        #only get integers
        for x in var:
            try:
                x = int(x)
            except:
                pass
            else:
                integers.append(x)
        for a in integers:
            inttemp.append(str(a))
        output = ''.join(inttemp)
        integer = int(output)
        return integer
    '''
    #Add colon for xx:xx if it is needed
    def colonorNot(data):
        temp=[]
        for x in data:
            temp.append(str(x))
        output = ''.join(temp)
        
        if len(output)==4:
            output = f"{output[:2]}:{output[2:]}"
        elif len(output)==3:
            output = f"{output[:1]}:{output[1:]}"
        else:
            return
        return output
     '''   
    #Detect whether the time is am or pm based off user input 
    def appendampm(x):
        if x is None:
            return
        if 'am' in x.lower():
            y = 'am'
        elif 'pm' in x.lower():
            y = 'pm'
        else:
            return None, x
        return y, str(x)
    '''       
    #Add zeros for time computation(1200, 1300, etc.)
    def zeros(data):
        print(f'{data}, {len(data)}, {type(data)}')
        if len(str(data))==1:
            slice = 1
            temp = list(str(data))
            temp.insert(slice, '00')
            data = int(''.join(temp))
            return data
        elif len(str(data))==2:
            slice = 2
            temp = list(str(data))
            temp.insert(slice, '00')
            data = int(''.join(temp))
            return data
        else:
            return(int(data))
    '''
    #Converts string into money format
    def money(string):
        if string[len(string)-2]=='.':
            string=f'{string}0'
        return f'${string}'
    
    #Fetching event date and time data
    def getData(position, dat):
        EventDate = str(dat.iloc[position][2])
        EventTime = str(dat.iloc[position][5])

        return EventDate, EventTime

    #Calculate all values
    def calculate(position, dat, c):

        #detecting whether each time period is am or pm
        dashindex = str(dat.iloc[position][4]).index("-")
        ampm = ['', '']
        ampm[0], Start = appendampm(str(dat.iloc[position][4][:dashindex]))
        ampm[1], End = appendampm(str(dat.iloc[position][4][dashindex:]))

        #idnetify vehicle type based on Excel Form
        vehicleTypes=[]
        if dat.iloc[position+c][6]:
            vehicleTypes.append(f'{dat.iloc[position+c][6]}')

        #Identify number of vehicles
        numVehicles = []
        if dat.iloc[position+c][5]:
            numVehicles.append(integerExtractor(str(dat.iloc[position+c][5])))

        numDrivers = dat.iloc[position][1]
        numEnforcementOfficers = dat.iloc[position][2]
        numSupervisors = dat.iloc[position][3]


        #Identify rate based on Excel Form
            # Opening JSON file
        with open('meta/menu.json') as json_file:
            RateDict = json.load(json_file)
        RateArr = []
        for vehicleType in vehicleTypes:
            RateArr.append(RateDict[vehicleType])

        timeStart, intStart = convert(dat, position, mod=0)
        timeEnd, intEnd = convert(dat, position, mod=1200)



        '''#Convert values:
        #xx00
        intStart = integerExtractor(Start)
        #xx00
        intEnd = integerExtractor(End)
        #xx00'''
        intRates = []
        for Rate in RateArr:
            intRates.append(float(Rate))
        '''#convert the start time into a computable number and convert to xx:xx
        intStart = zeros(str(intStart))
        timeStart = colonorNot(str(intStart))
        intEnd = zeros(str(intEnd))
        timeEnd = colonorNot(str(intEnd))'''

        Hours = 0

        #June 6th 23 - fixed 12 AM errors


        '''def ManageAM(time, ampm, mod):
            if time == 1200 and ampm == 'am':
                time=0+mod
            elif ampm=='pm' and time<1200:
                time+=1200

            return time

        intStart = ManageAM(intStart, ampm[0], 0)
        intEnd = ManageAM(intEnd, ampm[1], 1200)'''

        #intEnd is 2400 if 12 AM
        #Timing works on military time except for 12 AM

        if intEnd!=intStart:
            while intStart!=intEnd and Hours<24:
                Hours+=1
                intStart+=100

        #Dictionary template to fill 
        dictionaryArr = []
        TotalAmounts = []
        for i in range(len(intRates)): 
            totVehicles = round((numVehicles[i]*Hours)*intRates[i], 3)
            totSupervisors = round(float(RateDict['Supervisor(s)'])*Hours*numSupervisors, 3)
            totDrivers = round(numDrivers*float(RateDict['Motor Vehicle Operator(s)'])*Hours, 3)
            totEnforcementOfficers = round(numEnforcementOfficers*float(RateDict['Parking Enforcement Officers(s)'])*Hours, 3)
            VehicleDict = {
                "Description": f"{numVehicles[i]} {vehicleTypes[i]}, each for {Hours} hours",
                "Service Hours": f"{timeStart} {ampm[0]} - {timeEnd} {ampm[1]}",
                "Total Hours": f"{Hours*numVehicles[i]} hours",
                "Hourly Rate": f"{money(str(round(intRates[i], 3)))}",
                "Total Amount": f"{money(str(totVehicles))}"
            }

            dictionaryArr.append(VehicleDict)
        DriverDict = {
            "Description": f"{numDrivers} driver(s), each for {Hours} hours",
            "Service Hours": f"{timeStart} {ampm[0]} - {timeEnd} {ampm[1]}",
            "Total Hours": f"{Hours*numDrivers} hours",
            "Hourly Rate": f"{money(str(round(float(RateDict['Motor Vehicle Operator(s)']), 3)))}",
            "Total Amount": f"{money(str(totDrivers))} "
        }
        SupervisorDict = {
            "Description": f"{numSupervisors} supervisor(s) for {Hours} hours",
            "Service Hours": f"{timeStart} {ampm[0]} - {timeEnd} {ampm[1]}",
            "Total Hours": f"{Hours*numSupervisors} hours",
            "Hourly Rate": f"{money(str(round(float(RateDict['Supervisor(s)'])*numSupervisors, 3)))}",
            "Total Amount": f"{money(str(totSupervisors))} "
        }
        EnforcementOfficersDict = {
            "Description": f"{numEnforcementOfficers} enforcement officer(s) for {Hours} hours",
            "Service Hours": f"{timeStart} {ampm[0]} - {timeEnd} {ampm[1]}",
            "Total Hours": f"{Hours*numEnforcementOfficers} hours",
            "Hourly Rate": f"{money(str(round(float(RateDict['Parking Enforcement Officers(s)'])*numEnforcementOfficers, 3)))}",
            "Total Amount": f"{money(str(totEnforcementOfficers))} "
        }
        dictionaryArr.append(DriverDict)
        dictionaryArr.append(SupervisorDict)
        dictionaryArr.append(EnforcementOfficersDict)



        #df = pd.DataFrame(data = dict, index=[0])

        return dictionaryArr, round(totVehicles+totDrivers+totSupervisors+totEnforcementOfficers, 3)


    dfarr = []
    #Sum all costs and get a sum for the receipt
    totalvalues = []
    #Set up a final dataframe to append all dataframes to
    finaldf = pd.DataFrame(data=None)


    #n is used for figuring out what rows to delete

    n=0
    for a in swmaTimes:
        c=0
        print(a)
        while c<=2:
            print(c)
            print(f"swma.iloc[{a}+{c}][4]")
            print(type(swma.iloc[a+c][6]))
            print(swma.iloc[a+c][6])
            if type(swma.iloc[a][4]) is str and type(swma.iloc[a+c][6]) is str:
                dict0, value = calculate(a, swma, c)
                for i in dict0:
                    dfarr.append(i)
                    n=n+1
                totalvalues.append(value)
                print(dict0)
                c=c+1
            else:
                c=c+1

    for a in pemaTimes:
        c=0
        print(a)
        while c<=2:
            print(c)
            print(f"pema.iloc[{a}+{c}][4]")
            print(type(pema.iloc[a+c][6]))
            print(pema.iloc[a+c][6])
            if type(pema.iloc[a][4]) is str and type(pema.iloc[a+c][6]) is str:
                dict0, value = calculate(a, pema, c)
                for i in dict0:
                    dfarr.append(i)
                    n=n+1
                totalvalues.append(value)
                print(dict0)
                c=c+1
            else:
                c=c+1




    #Summation of cost
    values = round(sum(totalvalues), 3)

    dict1 = {
        "Description": '',
        "Service Hours": '',
        "Total Hours": '',
        "Hourly Rate": '',
        "Total Amount": f'{money(str(round(values, 2)))}'
    }
    #Fills the rest of the useless entries with dataframes of same size but no content
    dictfill = {
        "Description": '',
        "Service Hours": '',
        "Total Hours": '',
        "Hourly Rate": '',
        "Total Amount": ''
    }

    summation = pd.DataFrame(data = dict1, index=[0])
    dictfill = pd.DataFrame(data = dictfill, index=[0])

    while len(dfarr)<73:
        dfarr.append(dictfill)
    dfarr.append(summation)

    print(len(dfarr))

    #Copy is the template of the receipt that is used
    path = pd.read_excel('writeto.xlsx')   
    path = pd.DataFrame(path)

    #Always returns an error
    try:
        path.iloc[7][5] = getData(8, swma)[0]
    except Exception as e: 
        print(str(e))
        time.sleep(10)

        
    for i in range(19,93):  

        if str(dfarr[i-19]["Total Amount"])[0]!='0':
            path.iloc[i][1] = str(dfarr[i-19]["Description"])
            path.iloc[i][2] = str(dfarr[i-19]["Service Hours"])
            path.iloc[i][3] = str(dfarr[i-19]["Total Hours"])
            path.iloc[i][4] = str(dfarr[i-19]["Hourly Rate"])
            path.iloc[i][5] = str(dfarr[i-19]["Total Amount"])
        elif i==92:
            path.iloc[i][5] = str(dfarr[i-19]["Total Amount"]).strip('0').strip('\nName: Total Amount, dtype: object')
        else:
            path.iloc[i][1] = ''
            path.iloc[i][2] = ''
            path.iloc[i][3] = ''
            path.iloc[i][4] = ''
            path.iloc[i][5] = ''

    path.fillna(' ', inplace=True)
    path.replace('\n', '', regex=True)
    print(path)


    style_attrs = ["alignment", "border", "fill", "font", "number_format", "protection"]

    #Copy cell styling
    def cells(worksheet):
        """Return a generator for the sequence of cells in the worksheet"""
        for row in worksheet:
            for cell in row:
                yield cell


    def copy_attrs(src, dst, attrs=style_attrs):
        """Copy attributes from src to dst. Attributes are shallow-copied to avoid
        TypeError: unhashable type: 'StyleProxy'"""
        for name in attrs:
            setattr(dst, name, copy(getattr(src, name)))


    def copy_column_attrs(worksheet_src, worksheet_dst, attrs=style_attrs + ["width"]):
        """Copy ColumnDimension properties from worksheet_src to worksheet_dst.
        Only properties listed in attrs will be copied."""
        for column, dimensions in worksheet_src.column_dimensions.items():
            copy_attrs(
                src=dimensions,
                dst=worksheet_dst.column_dimensions[column],
                attrs=style_attrs + ["width"],
            )



    def copy_row_attrs(worksheet_src, worksheet_dst, attrs=style_attrs + ["height"]):
        """Copy RowDimension properties from worksheet_src to worksheet_dst.
        Only properties listed in attrs will be copied."""
        for row, dimensions in worksheet_src.row_dimensions.items():
            copy_attrs(
                src=dimensions,
                dst=worksheet_dst.row_dimensions[row],
                attrs=style_attrs + ["height"],
            )

    #write new data from df information
    def write_new_data(worksheet_dst, data):
        for i in range(19, 95):
            worksheet_dst.cell(row=i, column=1, value=data.iloc[i-2][0])
            worksheet_dst.cell(row=i, column=2, value=data.iloc[i-2][1])
            worksheet_dst.cell(row=i, column=3, value=data.iloc[i-2][2])
            worksheet_dst.cell(row=i, column=4, value=data.iloc[i-2][3])
            worksheet_dst.cell(row=i, column=5, value=data.iloc[i-2][4])
            worksheet_dst.cell(row=i, column=6, value=data.iloc[i-2][5])

    #we only copy cell styles here:
    def copy_cells(worksheet_src, worksheet_dst, data, attrs=style_attrs):
        num=0
        arr=[]
        """Copy cells from worksheet_src to worksheet_dst. If cells are styled
        then also copy the attributes listed in attrs."""
        cellvalues = cells(worksheet_src)
        for cell in cells(worksheet_src):
            arr.append(cell)
            cell_dst = worksheet_dst.cell(row=cell.row, column=cell.column)
            if cell.has_style:
                copy_attrs(cell, cell_dst, attrs=attrs)
            cell_dst.value = cell.value
            num=num+1
        write_new_data(worksheet_dst, data)

    #Clean end product before populating with data
    def delete_worksheet_cells(worksheet):
        worksheet.delete_cols(1, worksheet.max_column + 1)
        worksheet.delete_rows(1, worksheet.max_row + 1)

    #Template to copy style from
    wb_src = load_workbook("source.xlsx")
    ws_src = wb_src.active

    #Document to write to with new data
    wb_dst = load_workbook("writeto.xlsx")
    ws_dst = wb_dst.active

    delete_worksheet_cells(ws_dst)
    copy_column_attrs(ws_src, ws_dst)
    copy_row_attrs(ws_src, ws_dst)
    #copy_cells and the functions used inside of it are the main problems
    copy_cells(ws_src, ws_dst, path)


    #Auto fit columns
    dims = {}
    for row in ws_dst.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws_dst.column_dimensions[col].width = 30

    lis = []
    f=[]
    for col in range(48):
        lis.append(ws_dst.row_dimensions[col].height)
    for i in lis:
        if i is not None:
            f.append(i)
    ws_dst.row_dimensions[2].height=max(f)

    #n was referenced when adding dataframes
    ws_dst.delete_rows(21+n,94-(21+n))

    #Save as excel file

    class MyClass():
        def __init__(self, param):
            self.param = param
    
    def save_object(obj):
        try:
            with open("meta/data.pickle", "wb") as f:
                pickle.dump(obj, f, protocol=pickle.HIGHEST_PROTOCOL)
        except Exception as ex:
            print("Error during pickling object (Possibly unsupported):", ex)
            return None
    
    def load_object(filename):
        try:
            with open(filename, "rb") as f:
                return pickle.load(f)
        except Exception as ex:
            print("Error during unpickling object (Possibly unsupported):", ex)
            return None


    #Check whether .pickle contains data
    obj = int(load_object("meta/data.pickle"))
    if obj == None:
        save_object(1)
        wb_dst.save(f'Outputs/output{str(obj).zfill(4)}.xlsx')
        obj=obj+1
        save_object(obj)
    else:  
        wb_dst.save(f'Outputs/output{str(obj).zfill(4)}.xlsx')
        obj=obj+1
        save_object(obj)

if __name__ == '__main__':
    main()